"""
Extracts data for the RYB 2026 dropdown choropleth maps (see examples/RYB/2026/dropdown-maps.md).

For each dropdown map, reads one or more sheets from a chapter's "-for maps.xlsx" workbook and
writes a single JSON file consumed by examples/RYB/dropdown-choropleth.js.

Sheet layout (consistent across chapters, verified against CH02/CH04/CH08/CH13):
    A: NUTS code            C: Value
    I3/I4: report/chapter title   I6: map title   I7: map subtitle   I10: per-option label
    H18='Classes:' anchors a table (row count varies with number of classes):
        I: legend range label (unused - labels are regenerated from thresholds)
        K: color as "R#/G#/B#"
        N: threshold (present on all rows except the first, i.e. count = classes - 1)
    H='Footnotes:' / H='Sources:' anchor the note/source text in the next populated I cell.
Colors/thresholds are listed darkest-highest to lightest-lowest and are identical across every
option in a dropdown group (annotated "Fixed for Maps A-C" / "Hard coded" in the sheets), so they
are read once per group and reversed to ascending (low-to-high) order for eurostat-map.
"""

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]  # examples/RYB

# (xlsx path relative to ROOT, [(sheet, option code, option label override-or-None), ...], output json relative to ROOT)
GROUPS = [
    {
        "xlsx": "2026/CH02/data/RYB2026 CH02 Health-for maps.xlsx",
        "sheets": ["CH02M2A", "CH02M2B"],
        "out": "2026/CH02/data/CH02M02.json",
    },
    {
        "xlsx": "2026/CH04/data/RYB2026 CH04 Labour market-for maps.xlsx",
        "sheets": ["CH04M3A", "CH04M3B", "CH04M3C"],
        "out": "2026/CH04/data/CH04M03.json",
    },
    {
        "xlsx": "2026/CH08/data/RYB2026 CH08 Business-for maps.xlsx",
        "sheets": ["CH08M3A", "CH08M3B", "CH08M3C"],
        "out": "2026/CH08/data/CH08M03.json",
    },
    {
        "xlsx": "2026/CH08/data/RYB2026 CH08 Business-for maps.xlsx",
        "sheets": ["CH08M4A", "CH08M4B"],
        "out": "2026/CH08/data/CH08M04.json",
    },
    {
        "xlsx": "2026/CH13/data/RYB2026 CH13 Agriculture-for maps.xlsx",
        "sheets": ["CH13M2A", "CH13M2B", "CH13M2C"],
        "out": "2026/CH13/data/CH13M02.json",
    },
]

RGB_RE = re.compile(r"R(\d+)\s*/\s*G(\d+)\s*/\s*B(\d+)")


def clean_text(v):
    if v is None:
        return None
    return re.sub(r"\s+", " ", str(v)).strip()


def rgb_to_css(value):
    m = RGB_RE.search(str(value))
    if not m:
        raise ValueError(f"Unrecognised color format: {value!r}")
    r, g, b = (int(x) for x in m.groups())
    return f"rgb({r},{g},{b})"


def find_row_with_h_label(ws, label, max_row):
    for r in range(1, max_row + 1):
        if clean_text(ws.cell(row=r, column=8).value) == label:  # column H
            return r
    raise ValueError(f"Could not find row with H='{label}'")


def find_next_i_value(ws, start_row, max_lookahead=6):
    for r in range(start_row, start_row + max_lookahead):
        v = clean_text(ws.cell(row=r, column=9).value)  # column I
        if v:
            return v
    return None


def detect_nuts_level(ws, max_row):
    lengths = {}
    for r in range(2, max_row + 1):
        code = ws.cell(row=r, column=1).value
        if not code:
            continue
        lengths[len(str(code))] = lengths.get(len(str(code)), 0) + 1
    if not lengths:
        raise ValueError("Could not detect NUTS level: no codes found")
    mode_len = max(lengths, key=lengths.get)
    return max(mode_len - 2, 0)  # 2-char country code = level 0


def extract_classification(ws):
    max_row = ws.max_row
    classes_row = find_row_with_h_label(ws, "Classes:", max_row)

    raw_colors = []
    raw_thresholds = []
    r = classes_row
    while True:
        label = clean_text(ws.cell(row=r, column=9).value)  # column I
        if not label or label.lower().startswith("data not available"):
            break
        color_cell = ws.cell(row=r, column=11).value  # column K
        if color_cell is None:
            break
        raw_colors.append(rgb_to_css(color_cell))
        threshold_cell = ws.cell(row=r, column=14).value  # column N
        if threshold_cell is not None:
            raw_thresholds.append(float(threshold_cell))
        r += 1

    if len(raw_thresholds) != len(raw_colors) - 1:
        raise ValueError(f"Expected {len(raw_colors) - 1} thresholds for {len(raw_colors)} classes, got {len(raw_thresholds)}")

    # Sheet lists classes darkest/highest -> lightest/lowest; reverse to ascending (low -> high).
    colors = list(reversed(raw_colors))
    thresholds = list(reversed(raw_thresholds))

    decimals = 0
    for t in thresholds:
        frac = round(t - int(t), 6)
        if frac != 0:
            decimals = max(decimals, len(str(frac).split(".")[-1].rstrip("0")) or 1)

    footnote = find_next_i_value(ws, find_row_with_h_label(ws, "Footnotes:", max_row))
    source = find_next_i_value(ws, find_row_with_h_label(ws, "Sources:", max_row))

    return {
        "colors": colors,
        "thresholds": thresholds,
        "decimals": decimals,
        "footnote": footnote,
        "source": source,
    }


def fallback_label(code, title, subtitle):
    # Some sheets differentiate options only by year in the subtitle (e.g. CH13 map2 A/B/C),
    # with I10 left blank. Prefer that year over a bare letter code when present.
    m = re.search(r"\b(19|20)\d{2}\b", subtitle or "")
    if m:
        return m.group(0)
    return code


def extract_option(ws, code):
    max_row = ws.max_row
    data = {}
    for r in range(2, max_row + 1):
        nuts = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=3).value
        if not nuts or value in (None, ":", ""):
            continue
        try:
            data[str(nuts)] = float(value)
        except (TypeError, ValueError):
            continue

    title = clean_text(ws.cell(row=6, column=9).value)  # I6
    subtitle = clean_text(ws.cell(row=7, column=9).value)  # I7
    label = clean_text(ws.cell(row=10, column=9).value) or fallback_label(code, title, subtitle)  # I10

    return {
        "code": code,
        "label": label,
        "title": title,
        "subtitle": subtitle,
        "chapterTitle": clean_text(ws.cell(row=4, column=9).value),  # I4
        "unitText": clean_text(ws.cell(row=17, column=9).value) or "",  # I17
        "data": data,
    }


def extract_group(xlsx_path, sheet_names):
    wb = load_workbook(xlsx_path, data_only=True)

    options = []
    classification = None
    nuts_level = None

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        code = sheet_name[-1]  # trailing letter, e.g. 'A'
        options.append(extract_option(ws, code))

        sheet_classification = extract_classification(ws)
        sheet_nuts_level = detect_nuts_level(ws, ws.max_row)

        if classification is None:
            classification = sheet_classification
            nuts_level = sheet_nuts_level
        else:
            if sheet_classification["colors"] != classification["colors"] or sheet_classification["thresholds"] != classification["thresholds"]:
                raise ValueError(f"Classification mismatch between options in {xlsx_path.name} ({sheet_names})")
            if sheet_nuts_level != nuts_level:
                raise ValueError(f"NUTS level mismatch between options in {xlsx_path.name} ({sheet_names})")

    unit_text = options[0]["unitText"]
    legend_title = unit_text

    return {
        "nutsLevel": nuts_level,
        "scale": "20M",
        "unitText": unit_text,
        "legendTitle": legend_title,
        "colors": classification["colors"],
        "thresholds": classification["thresholds"],
        "decimals": classification["decimals"],
        "options": [
            {
                "code": o["code"],
                "label": o["label"],
                "title": o["title"],
                "subtitle": o["subtitle"],
                "chapterTitle": o["chapterTitle"],
                "footnote": classification["footnote"],
                "source": classification["source"],
                "data": o["data"],
            }
            for o in options
        ],
    }


def main():
    for group in GROUPS:
        xlsx_path = ROOT / group["xlsx"]
        out_path = ROOT / group["out"]
        print(f"Extracting {out_path.relative_to(ROOT)} from {xlsx_path.name} {group['sheets']}...")

        result = extract_group(xlsx_path, group["sheets"])

        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        n_regions = len(result["options"][0]["data"])
        print(f"  nutsLevel={result['nutsLevel']} classes={len(result['colors'])} decimals={result['decimals']} regions~={n_regions}")

    print("Done.")


if __name__ == "__main__":
    main()
