"""Extract URE 2026 GDP-change map data from Economy_URE2026.xlsx sheet M1/M01."""

import json
import re
from pathlib import Path

from openpyxl import load_workbook

URE_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = URE_ROOT / "2026/data/input/Economy_URE2026.xlsx"
OUTPUT = URE_ROOT / "2026/data/gdp-change.json"
RGB_RE = re.compile(r"R(\d+)\s*/\s*G(\d+)\s*/\s*B(\d+)")


def clean_text(value):
    return re.sub(r"\s+", " ", str(value)).strip() if value is not None else None


def rgb_to_css(value):
    match = RGB_RE.fullmatch(clean_text(value) or "")
    if not match:
        raise ValueError(f"Unrecognised color format: {value!r}")
    return f"rgb({','.join(match.groups())})"


def json_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip() == ":":
        return ":"
    try:
        return float(value)
    except (TypeError, ValueError):
        return str(value).strip()


def next_i_value(ws, h_label):
    for row in range(1, ws.max_row + 1):
        if clean_text(ws.cell(row, 8).value) == h_label:
            for candidate in range(row, min(row + 7, ws.max_row + 1)):
                value = clean_text(ws.cell(candidate, 9).value)
                if value:
                    return value
    raise ValueError(f"Could not find metadata following H='{h_label}'")


def extract():
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheet_name = next((name for name in ("M1", "M01") if name in workbook.sheetnames), None)
    if not sheet_name:
        raise ValueError(f"Could not find sheet M1/M01 in {WORKBOOK.name}")
    ws = workbook[sheet_name]

    gdp_change = {}
    urban_rural_type = {}
    for nuts, _, gdp, urban_rural in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        nuts = clean_text(nuts)
        if not nuts:
            break
        gdp = json_value(gdp)
        urban_rural = json_value(urban_rural)
        if gdp is not None:
            gdp_change[nuts] = gdp
        if urban_rural is not None:
            urban_rural_type[nuts] = urban_rural

    colors_by_class = {}
    for row in range(12, 22):
        class_id = clean_text(ws.cell(row, 10).value)
        color = ws.cell(row, 11).value
        if class_id and color:
            colors_by_class[class_id] = rgb_to_css(color)

    return {
        "sheet": ws.title,
        "nutsLevel": "mixed",
        "nutsYear": 2024,
        "scale": "60M",
        "chapterTitle": clean_text(ws["I4"].value),
        "title": clean_text(ws["I6"].value),
        "subtitle": clean_text(ws["I7"].value),
        "gdpLegendLabel": clean_text(ws["I10"].value),
        "urbanRuralLegendLabel": clean_text(ws["M12"].value),
        "thresholds": sorted(float(ws.cell(row, 19).value) for row in (14, 16)),
        "urbanRuralLabels": {
            str(int(ws.cell(row, 21).value)): clean_text(ws.cell(row, 20).value)
            for row in range(14, 17)
        },
        "colorsByClass": colors_by_class,
        "classMatrix": [[int(ws.cell(row, col).value) for col in range(13, 16)] for row in range(15, 18)],
        "footnote": next_i_value(ws, "Footnotes:"),
        "source": next_i_value(ws, "Sources:"),
        "stats": {"gdpChange": gdp_change, "urbanRuralType": urban_rural_type},
    }


def main():
    result = extract()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8") as output_file:
        json.dump(result, output_file, ensure_ascii=False, indent=2)
    print(
        f"Wrote {OUTPUT.relative_to(URE_ROOT)} from {result['sheet']}: "
        f"{len(result['stats']['gdpChange'])} GDP values and "
        f"{len(result['stats']['urbanRuralType'])} urban-rural values"
    )


if __name__ == "__main__":
    main()
